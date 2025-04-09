from datetime import datetime
import re

from flask_login import current_user

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from webapp.home import exceptions as exceptions
from webapp.home.home_utils import get_check_metadata_status
import webapp.home.utils.load_and_save as load_and_save
from webapp.home.utils.load_and_save import load_eml, save_eml
from webapp.home.utils.lists import compose_attribute_mscale
from webapp.home.utils.node_utils import new_child_node
from metapype.model.node import Node
from metapype.eml import names

standard_units = ['', 'acre', 'ampere', 'amperePerMeter', 'amperePerMeterSquared', 'angstrom', 'are', 'atmosphere',
                  'bar', 'becquerel', 'britishThermalUnit', 'bushel', 'bushelPerAcre', 'calorie', 'candela',
                  'candelaPerMeterSquared', 'celsius', 'centigram', 'centimeter', 'centimeterCubed',
                  'centimeterPerSecond', 'centimeterSquared', 'centisecond', 'coulomb', 'decibar', 'decigram',
                  'decimeter', 'decisecond', 'degree', 'dekagram', 'dekameter', 'dekasecond', 'dimensionless',
                  'equivalentPerLiter', 'fahrenheit', 'farad', 'fathom', 'foot', 'Foot_Gold_Coast', 'Foot_US',
                  'footCubedPerSecond', 'footPerDay', 'footPerHour', 'footPerSecond', 'footPound', 'footSquared',
                  'footSquaredPerDay', 'gallon', 'grad', 'gram', 'gramPerCentimeterCubed', 'gramPercentimeterSquared',
                  'gramPerCentimeterSquaredPerSecond', 'gramPerDayPerHectare', 'gramPerDayPerLiter', 'gramPerGram',
                  'gramPerLiter', 'gramPerMeterSquared', 'gramPerMeterSquaredPerDay', 'gramPerMeterSquaredPerYear',
                  'gramPerMilliliter', 'gramPerYear', 'gray', 'hectare', 'hectogram', 'hectometer', 'hectopascal',
                  'hectosecond', 'henry', 'hertz', 'hour', 'inch', 'inchCubed', 'inchPerHour', 'inverseCentimeter',
                  'inverseMeter', 'joule', 'katal', 'kelvin', 'kilogram', 'kilogramPerHectare',
                  'kilogramPerHectarePerYear', 'kilogramPerMeterCubed', 'kilogramPerMeterSquared',
                  'kilogramPerMeterSquaredPerDay', 'kilogramPerMeterSquaredPerSecond', 'kilogramPerMeterSquaredPerYear',
                  'kilogramPerSecond', 'kilohertz', 'kiloliter', 'kilometer', 'kilometerPerHour', 'kilometerSquared',
                  'kilopascal', 'kilosecond', 'kilovolt', 'kilowatt', 'kilowattPerMeterSquared', 'knot', 'langley',
                  'langleyPerDay', 'Link_Clarke', 'liter', 'literPerHectare', 'literPerLiter', 'literPerMeterSquared',
                  'literPerSecond', 'lumen', 'lux', 'megagram', 'megagramPerMeterCubed', 'megahertz',
                  'megajoulePerMeterSquaredPerDay', 'megameter', 'megapascal', 'megasecond', 'megavolt', 'megawatt',
                  'meter', 'meterCubed', 'meterCubedPerHectare', 'meterCubedPerKilogram', 'meterCubedPerMeterCubed',
                  'meterCubedPerMeterSquared', 'meterCubedPerSecond', 'meterPerDay', 'meterPerGram', 'meterPerSecond',
                  'meterPerSecondSquared', 'meterSquared', 'meterSquaredPerDay', 'meterSquaredPerHectare',
                  'meterSquaredPerKilogram', 'meterSquaredPerSecond', 'microequivalentPerLiter', 'microgram',
                  'microgramPerGram', 'microgramPerGramPerDay', 'microgramPerGramPerHour', 'microgramPerGramPerWeek',
                  'microgramPerLiter', 'microliter', 'microliterPerLiter', 'micrometer', 'micrometerCubedPerGram',
                  'micromolePerCentimeterSquaredPerSecond', 'micromolePerGram', 'micromolePerGramPerDay',
                  'micromolePerGramPerHour', 'micromolePerGramPerSecond', 'micromolePerKilogram', 'micromolePerLiter',
                  'micromolePerMeterSquaredPerSecond', 'micromolePerMole', 'microsecond',
                  'microwattPerCentimeterSquaredPerNanometer', 'microwattPerCentimeterSquaredPerNanometerPerSteradian',
                  'microwattPerCentimeterSquaredPerSteradian', 'mile', 'milePerHour', 'milePerMinute', 'milePerSecond',
                  'mileSquared', 'millibar', 'milliequivalentPerLiter', 'milligram', 'milligramPerKilogram',
                  'milligramPerLiter', 'milligramPerMeterCubed', 'milligramPerMeterCubedPerDay',
                  'milligramPerMeterSquared', 'milligramPerMeterSquaredPerDay', 'milligramPerMilliliter', 'millihertz',
                  'milliliter', 'milliliterPerLiter', 'millimeter', 'millimeterPerDay', 'millimeterPerSecond',
                  'millimeterSquared', 'millimolePerGram', 'millimolePerKilogram', 'millimolePerLiter',
                  'millimolePerMeterCubed', 'millimolePerMole', 'millisecond', 'millivolt', 'milliwatt', 'minute',
                  'mole', 'molePerGram', 'molePerKilogram', 'molePerKilogram', 'molePerKilogramPerSecond',
                  'molePerLiter', 'molePerMeterCubed', 'molePerMeterSquaredPerSecond', 'molePerMole', 'nanogram',
                  'nanogramPerGram', 'nanogramPerGramPerHour', 'nanoliterPerLiter', 'nanometer',
                  'nanomolePerGramPerDay', 'nanomolePerGramPerHour', 'nanomolePerGramPerSecond', 'nanomolePerKilogram',
                  'nanomolePerLiter', 'nanomolePerMole', 'nanosecond', 'nauticalMile', 'newton', 'nominalDay',
                  'nominalHour', 'nominalLeapYear', 'nominalMinute', 'nominalWeek', 'nominalYear', 'number',
                  'numberPerGram', 'numberPerHectare', 'numberPerKilometerSquared', 'numberPerLiter',
                  'numberPerMeterCubed', 'numberPerMeterSquared', 'numberPerMilliliter', 'ohm', 'ohmMeter', 'pascal',
                  'percent', 'permil', 'pint', 'pound', 'poundPerAcre', 'poundPerInchSquared', 'quart', 'radian',
                  'second', 'siemens', 'siemensPerCentimeter', 'siemensPerMeter', 'sievert', 'steradian', 'tesla',
                  'ton', 'tonne', 'tonnePerHectare', 'tonnePerYear', 'volt', 'watt', 'wattPerMeterSquared',
                  'wattPerMeterSquaredPerNanometer', 'wattPerMeterSquaredPerNanometerPerSteradian',
                  'wattPerMeterSquaredPerSteradian', 'weber', 'yard', 'Yard_Indian', 'yardPerSecond', 'yardSquared']

# Globals
wb = None
ws = None
# Row and column indices are 1-based
row = None
col = None
maxcol = None
maxrow = None
data_validation = None

def rgb_to_hex(rgb):
    return '{:02X}{:02X}{:02X}'.format(*rgb)

REQUIRED = rgb_to_hex((232, 200, 200))
RECOMMENDED = rgb_to_hex((252, 235, 166))
OPTIONAL = rgb_to_hex((211, 237, 244))
RARELY_USED = rgb_to_hex((242, 242, 242))
FONT_SIZE = 12


def next_spreadsheet_column(column, i=1):
    global maxcol
    """
    Given a spreadsheet column name (e.g., 'A', 'Z', 'AA') and an integer i,
    returns the column name i positions greater.
    """
    # Convert the current column to a number, add i, and convert back
    current_col_num = column_index_from_string(column)
    next_col_num = current_col_num + i
    if maxcol is None or next_col_num > maxcol:
        maxcol = next_col_num
    return get_column_letter(next_col_num)


def is_categorical(attribute_node):
    enumerated_domain_node = attribute_node.find_descendant(names.ENUMERATEDDOMAIN)
    if enumerated_domain_node:
        return True
    return False


def is_datetime(attribute_node):
    datetime_node = attribute_node.find_descendant(names.DATETIME)
    if datetime_node:
        return True
    return False


def interval_or_ratio_node(attribute_node):
    measurement_scale_node = attribute_node.find_child(names.MEASUREMENTSCALE)
    if measurement_scale_node:
        interval_node = measurement_scale_node.find_child(names.INTERVAL)
        if interval_node:
            return interval_node
        ratio_node = measurement_scale_node.find_child(names.RATIO)
        if ratio_node:
            return ratio_node
    return None


def is_numerical(attribute_node):
    if interval_or_ratio_node(attribute_node):
        return True
    return False


####################################################################################################
# Functions for writing metadata to a spreadsheet (i.e., do download)
####################################################################################################

def extract_column_row(cell_str):
    # Use regular expression to match the column letters and row numbers
    match = re.match(r"([A-Z]+)([0-9]+)", cell_str, re.I)
    if match:
        # Extract the column and row parts
        column, row = match.groups()
        return column.upper(), int(row)  # Return column as uppercase and row as integer
    else:
        return None, None  # In case the input does not match the expected format


def set_maxcolrow(cell):
    global maxcol, maxrow
    col, row = extract_column_row(cell)
    col_num = column_index_from_string(col)
    if maxcol is None or col_num > maxcol:
        maxcol = col_num
    if maxrow is None or row > maxrow:
        maxrow = row


def highlight_cell(cell, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    ws[cell].font = Font(size=FONT_SIZE)
    ws[cell].fill = fill
    ws[cell].border = border
    set_maxcolrow(cell)


def bold(cell, text):
    ws[cell].font = Font(size=FONT_SIZE, bold=True)
    ws[cell] = text
    set_maxcolrow(cell)


def italic(cell, text):
    ws[cell].font = Font(size=FONT_SIZE, italic=True)
    ws[cell] = text
    set_maxcolrow(cell)


def normal(cell, text):
    ws[cell].font = Font(size=FONT_SIZE)
    ws[cell] = text
    set_maxcolrow(cell)


def required(cell, text):
    normal(cell, text)
    highlight_cell(cell, REQUIRED)


def recommended(cell, text):
    normal(cell, text)
    highlight_cell(cell, RECOMMENDED)


def optional(cell, text):
    normal(cell, text)
    highlight_cell(cell, OPTIONAL)


def rarely_used(cell, text):
    normal(cell, text)
    highlight_cell(cell, RARELY_USED)


def create_data_validation():
    global data_validation, ws, wb
    ws = wb.active
    # We need to reference the standard units list on Sheet2 in this way because the list is limited to 255 characters
    #  if specified directly in the formula1 argument of the DataValidation object.
    formula1 = "'Standard Units'!$A$2:$A$" + str(len(standard_units) + 1)
    data_validation = DataValidation(type="list", formula1=formula1, allow_blank=True)
    data_validation.error = 'Your entry is not an allowed standard unit. Please choose from the drop-down list or use a custom unit.'
    data_validation.errorTitle = 'Not an allowed standard unit'
    ws.add_data_validation(data_validation)


def start_table(package_name, table_name, description=''):
    global row
    bold('A1', 'Package Name')
    normal('B1', package_name)
    bold('A2', 'Table Name')
    normal('B2', table_name)
    bold('A3', 'Date Downloaded')
    normal('B3', datetime.now().strftime('%Y-%m-%d'))
    row = 4


def insert_color_key():
    bold('D1', 'COLOR KEY')
    normal('D2', 'Required')
    normal('D3', 'Recommended')
    required('E2', '')
    recommended('E3', '')
    normal('L1', 'Spreadsheet format: 001')


def set_column_widths():
    ws.column_dimensions['A'].width = 30
    for i in range(maxcol):
        ws.column_dimensions[next_spreadsheet_column('A', i + 1)].width = 30
        for j in range(maxrow):
            ws.cell(row=j+1, column=i+1).number_format = '@'

    for i in range(maxrow):
        for j in range(maxcol):
            is_bold = ws.cell(i + 1, j + 1).font.bold
            is_italic = ws.cell(i + 1, j + 1).font.italic
            color = ws.cell(i + 1, j + 1).font.color
            ws.cell(i + 1, j + 1).font = Font(size=FONT_SIZE, bold=is_bold, italic=is_italic, color=color)


def remove_sheet_protection(first_row, last_row, first_col, last_col):
    return
    for row in range(first_row, last_row+1):
        for col in range(first_col, last_col+1):
            ws.cell(row=row, column=col).protection = Protection(locked=False)
            ws.cell(row=row, column=col).font = Font(size=FONT_SIZE)


def create_sheet2():
    global wb
    ws2 = wb.create_sheet('Standard Units')
    ws2['A1'] = 'Allowed Standard Units:'
    ws2['A1'].font = Font(size=FONT_SIZE, bold=True)
    row = 2
    for unit in standard_units:
        ws2[f'A{row}'] = unit
        ws2[f'A{row}'].font = Font(size=FONT_SIZE)
        row += 1


def create_sheet():
    global wb, ws, row
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    # ws.protection.sheet = True
    # ws.protection.formatColumns = True
    # ws.protection.password = 'ezEML'
    ws.freeze_panes = 'B1'
    row = 1

    insert_color_key()


def add_all_columns_headers():
    global row
    bold('A6', 'ALL COLUMNS')
    row = 7
    bold(f'A{str(row)}', 'Column')
    bold(f'B{str(row)}', 'Type')
    bold(f'C{str(row)}', 'Definition')
    bold(f'D{str(row)}', 'Label')
    bold(f'E{str(row)}', 'Storage Type')
    bold(f'F{str(row)}', 'Storage Type System')

    bold(f'G{str(row - 1)}', 'MISSING VALUE CODES')
    bold(f'G{str(row)}', 'MVC 1')
    bold(f'H{str(row)}', 'MVC 1 Explanation')
    bold(f'I{str(row)}', 'MVC 2')
    bold(f'J{str(row)}', 'MVC 2 Explanation')
    bold(f'K{str(row)}', 'MVC 3')
    bold(f'L{str(row)}', 'MVC 3 Explanation')


def add_numerical_columns_headers():
    global row
    bold(f'A{str(row)}', 'NUMERICAL')
    ws[f'C{str(row)}'].font = Font(size=FONT_SIZE, bold=True, italic=True, color='4682B4')
    ws[f'C{str(row)}'] = 'Either a Standard Unit or Custom Unit is required.'
    row += 1
    bold(f'A{str(row)}', 'Column')
    bold(f'B{str(row)}', 'Number Type')
    bold(f'C{str(row)}', 'Standard Unit')
    bold(f'D{str(row)}', 'Custom Unit')
    bold(f'E{str(row)}', 'Custom Unit Description')
    bold(f'F{str(row)}', 'Precision')
    bold(f'G{str(row)}', 'Bounds Minimum')
    bold(f'H{str(row)}', 'Bounds Maximum')


def add_datetime_columns_headers():
    global row
    bold(f'A{str(row)}', 'DATETIME')
    row += 1
    bold(f'A{str(row)}', 'Column')
    bold(f'B{str(row)}', 'Format String')
    bold(f'C{str(row)}', 'Precision')
    bold(f'D{str(row)}', 'Bounds Minimum')
    bold(f'E{str(row)}', 'Bounds Maximum')


def add_categorical_columns_headers():
    global row
    bold(f'A{str(row)}', 'CATEGORICAL')
    row += 1
    # bold(f'A{str(row)}', 'Column')
    # bold(f'B{str(row)}', 'Code')
    # bold(f'C{str(row)}', 'Definition')


def add_column(column, header, values, color=normal):
    global row
    bold(column + str(row), header)
    row += 1
    for value in values:
        color(column + str(row), value)
        row += 1


def add_column_names(names):
    global row
    row = 7
    add_column('A', 'Column', names)


def add_column_types(types):
    global row
    row = 7
    add_column('B', 'Type', types)


def add_column_definitions(definitions):
    global row
    row = 7
    add_column('C', 'Definition', definitions, recommended)


def add_column_labels(labels):
    global row
    row = 7
    add_column('D', 'Label', labels)


def add_storage_types(storage_types):
    global row
    row = 7
    add_column('E', 'Storage Type', storage_types)


def add_storage_type_systems(storage_type_systems):
    global row
    row = 7
    add_column('F', 'Storage Type System', storage_type_systems)


def add_missing_value_codes(attribute_nodes):
    def add_missing_value_code(missing_value_node, column):
        code_node = missing_value_node.find_child(names.CODE)
        explanation_node = missing_value_node.find_child(names.CODEEXPLANATION)
        if code_node and code_node.content:
            normal(column + str(row), code_node.content)
            required(next_spreadsheet_column(column) + str(row), explanation_node.content if explanation_node else '')

    global row
    row = 8
    for attribute_node in attribute_nodes:
        missing_value_nodes = attribute_node.find_all_children(names.MISSINGVALUECODE)
        i = 0
        for missing_value_node in missing_value_nodes:
            add_missing_value_code(missing_value_node, next_spreadsheet_column('G', 2 * i))
            i += 1
            if i > 2:
                break
        row += 1


def add_numerical_columns(attribute_nodes, eml_node):
    global row
    def get_custom_unit_description(eml_node, custom_unit_name):
        # get description, if any, from an additionalMetadata section
        additional_metadata_nodes = eml_node.find_all_children(names.ADDITIONALMETADATA)
        for additional_metadata_node in additional_metadata_nodes:
            metadata_node = additional_metadata_node.find_child(names.METADATA)
            if metadata_node:
                unit_list_node = metadata_node.find_child(names.UNITLIST)
                if unit_list_node:
                    unit_nodes = unit_list_node.find_all_children(names.UNIT)
                    unit_node = None
                    for node in unit_nodes:
                        if node.attribute_value('name') == custom_unit_name:
                            unit_node = node
                            break
                    if unit_node:
                        description_node = unit_node.find_child(names.DESCRIPTION)
                        if description_node:
                            return description_node.content
        return ''

    def add_numerical_column(attribute_node, eml_node):
        global row, data_validation
        if not is_numerical(attribute_node):
            return
        row += 1
        remove_sheet_protection(row, row, 2, 8)
        data_validation.add(f'C{str(row)}')

        normal('A' + str(row), attribute_node.find_child(names.ATTRIBUTENAME).content)
        number_type_node = attribute_node.find_descendant(names.NUMBERTYPE)
        if number_type_node:
            required('B' + str(row), number_type_node.content)

        measurement_scale_node = attribute_node.find_child(names.MEASUREMENTSCALE)
        interval_or_ratio_node = measurement_scale_node.find_child(names.INTERVAL) or measurement_scale_node.find_child(names.RATIO)
        if interval_or_ratio_node:
            unit_node = interval_or_ratio_node.find_child(names.UNIT)
            if unit_node:
                standard_unit_node = unit_node.find_child(names.STANDARDUNIT)
                custom_unit_node = unit_node.find_child(names.CUSTOMUNIT)
                if custom_unit_node:
                    required('D' + str(row), custom_unit_node.content)
                    description = get_custom_unit_description(eml_node, custom_unit_node.content)
                    recommended('E' + str(row), description)
                elif standard_unit_node:
                    required('C' + str(row), standard_unit_node.content)
                else:
                    required('C' + str(row), '')
                    normal('D' + str(row), '')
            else:
                required('C' + str(row), '')
                normal('D' + str(row), '')
            precision_node = interval_or_ratio_node.find_child(names.PRECISION)
            if precision_node:
                normal('F' + str(row), precision_node.content)
            minimum_node = interval_or_ratio_node.find_descendant(names.MINIMUM)
            if minimum_node:
                normal('G' + str(row), minimum_node.content)
            maximum_node = interval_or_ratio_node.find_descendant(names.MAXIMUM)
            if maximum_node:
                normal('H' + str(row), maximum_node.content)

    for attribute_node in attribute_nodes:
        add_numerical_column(attribute_node, eml_node)


def add_datetime_columns(attribute_nodes):
    global row
    def add_datetime_column(attribute_node):
        global row
        if not is_datetime(attribute_node):
            return
        row += 1
        remove_sheet_protection(row, row, 2, 5)
        normal('A' + str(row), attribute_node.find_child(names.ATTRIBUTENAME).content)
        format_node = attribute_node.find_single_node_by_path([names.MEASUREMENTSCALE, names.DATETIME, names.FORMATSTRING])
        required('B' + str(row), '')
        if format_node:
            required('B' + str(row), format_node.content)
        precision_node = attribute_node.find_descendant(names.PRECISION)
        if precision_node:
            normal('C' + str(row), precision_node.content)
        minimum_node = attribute_node.find_descendant(names.MINIMUM)
        if minimum_node:
            normal('D' + str(row), minimum_node.content)
        maximum_node = attribute_node.find_descendant(names.MAXIMUM)
        if maximum_node:
            normal('E' + str(row), maximum_node.content)

    for attribute_node in attribute_nodes:
        add_datetime_column(attribute_node)


def add_categorical_columns(attribute_nodes, header_row):
    global row, col

    def add_categorical_column(attribute_node, start_row):
        global row, col
        if not is_categorical(attribute_node):
            return
        row = start_row
        bold(f'{col}{row}', f'Column: {attribute_node.find_child(names.ATTRIBUTENAME).content}')
        row += 1
        # remove_sheet_protection(row, row + 100, col, col + 1)
        code_definition_nodes = []
        codes = []
        definitions = []
        attribute_node.find_all_descendants(names.CODEDEFINITION, code_definition_nodes)
        for code_definition_node in code_definition_nodes:
            code_node = code_definition_node.find_child(names.CODE)
            definition_node = code_definition_node.find_child(names.DEFINITION)
            if code_node:
                codes.append(code_node.content)
            else:
                codes.append('')
            if definition_node:
                definitions.append(definition_node.content)
            else:
                definitions.append('')
        row = start_row + 1
        add_column(col, "Code", codes)
        row = start_row + 1
        add_column(next_spreadsheet_column(col), "Definition", definitions, required)
        col = next_spreadsheet_column(col, 3)

    start_row = row
    col = 'A'
    for attribute_node in attribute_nodes:
        add_categorical_column(attribute_node, start_row)


def generate_data_entry_spreadsheet(data_table_node, filename, data_table_name):
    global row
    if not data_table_node or not data_table_node.name:
        return None

    eml_node = load_eml(filename)

    create_sheet()
    add_all_columns_headers()
    create_sheet2()
    create_data_validation()

    entity_name_node = data_table_node.find_child(names.ENTITYNAME)
    entity_name = entity_name_node.content if entity_name_node else ''
    if not entity_name:
        msg = f'No entity name found for data table {data_table_name}'
        raise exceptions.DataTableNameNotFound(msg)

    start_table(filename, entity_name_node.content if entity_name_node else '')

    attribute_nodes = data_table_node.find_child(names.ATTRIBUTELIST).children
    column_names = []
    column_types = []
    column_definitions = []
    column_labels = []
    storage_types = []
    storage_type_systems = []
    for attribute_node in attribute_nodes:
        attribute_name_node = attribute_node.find_child(names.ATTRIBUTENAME)
        column_names.append(attribute_name_node.content if attribute_name_node else '')
        column_types.append(compose_attribute_mscale(attribute_node))
        column_definitions.append(attribute_node.find_child(names.ATTRIBUTEDEFINITION).content if attribute_node.find_child(names.ATTRIBUTEDEFINITION) else '')
        column_labels.append(attribute_node.find_child(names.ATTRIBUTELABEL).content if attribute_node.find_child(names.ATTRIBUTELABEL) else '')
        storage_type_node = attribute_node.find_child(names.STORAGETYPE)
        if storage_type_node:
            storage_types.append(storage_type_node.content)
            # storage_type_systems.append(storage_type_node.attributes.get(names.TYPESYSTEM))
            storage_type_systems.append(storage_type_node.attributes.get('typeSystem')) # TODO
        else:
            storage_types.append('')
            storage_type_systems.append('')

    add_column_names(column_names)
    add_column_types(column_types)
    add_column_definitions(column_definitions)
    add_column_labels(column_labels)
    add_storage_types(storage_types)
    add_storage_type_systems(storage_type_systems)
    add_missing_value_codes(attribute_nodes)
    remove_sheet_protection(8, 8 + len(column_names) - 1, 3, 13)

    if 'Numerical' in column_types:
        row += 2
        add_numerical_columns_headers()
        add_numerical_columns(attribute_nodes, eml_node)
        row += 1

    if 'DateTime' in column_types:
        row += 2
        add_datetime_columns_headers()
        add_datetime_columns(attribute_nodes)
        row += 1

    if 'Categorical' in column_types:
        row += 2
        add_categorical_columns_headers()
        add_categorical_columns(attribute_nodes, row)

    set_column_widths()

    import os
    from pathlib import Path
    import webapp.auth.user_data as user_data
    user_folder = user_data.get_user_folder_name()
    sheets_folder = os.path.join(user_folder, 'spreadsheets')
    Path(sheets_folder).mkdir(parents=True, exist_ok=True)
    outfile = os.path.join(sheets_folder, f'{filename}__{data_table_name}.xlsx')
    wb.save(outfile)
    return outfile


####################################################################################################
# Functions for reading metadata from a spreadsheet (i.e., do upload)
####################################################################################################

def get_package_name(sheet):
    return sheet['B1'].value


def get_data_table_name(sheet):
    return sheet['B2'].value


def get_column_values(sheet, column, row=8, length=None):
    values = []
    i = row
    while True:
        cell = sheet[f'{column}{i}']
        i += 1
        if not cell.value and not length:
            break
        values.append(cell.value)
        if length and i - row >= length:
            break
    return values


def get_column_names(sheet):
    return get_column_values(sheet, 'A')


def get_column_types(sheet, length=None):
    return get_column_values(sheet, 'B', length=length)


def get_column_definitions(sheet, length=None):
    return get_column_values(sheet, 'C', length=length)


def get_column_labels(sheet, length=None):
    return get_column_values(sheet, 'D', length=length)


def get_storage_types(sheet, length=None):
    return get_column_values(sheet, 'E', length=length)


def get_storage_type_systems(sheet, length=None):
    return get_column_values(sheet, 'F', length=length)


def get_missing_values(sheet, length=None):
    mvc1 = get_column_values(sheet, 'G', length=length)
    mvc_explanations_1 = get_column_values(sheet, 'H', length=length)
    mvc2 = get_column_values(sheet, 'I', length=length)
    mvc_explanations_2 = get_column_values(sheet, 'J', length=length)
    mvc3 = get_column_values(sheet, 'K', length=length)
    mvc_explanations_3 = get_column_values(sheet, 'L', length=length)
    return mvc1, mvc_explanations_1, mvc2, mvc_explanations_2, mvc3, mvc_explanations_3


def get_numerical_variables(sheet, num_numerical_variables):
    global row
    number_types = get_column_values(sheet, 'B', row=row, length=num_numerical_variables)
    standard_units = get_column_values(sheet, 'C', row=row, length=num_numerical_variables)
    custom_units = get_column_values(sheet, 'D', row=row, length=num_numerical_variables)
    custom_unit_descriptions = get_column_values(sheet, 'E', row=row, length=num_numerical_variables)
    precisions = get_column_values(sheet, 'F', row=row, length=num_numerical_variables)
    bounds_minima = get_column_values(sheet, 'G', row=row, length=num_numerical_variables)
    bounds_maxima = get_column_values(sheet, 'H', row=row, length=num_numerical_variables)
    return number_types, standard_units, custom_units, custom_unit_descriptions, precisions, bounds_minima, bounds_maxima


def get_datetime_variables(sheet, num_datetime_variables):
    format_strings = get_column_values(sheet, 'B', row=row, length=num_datetime_variables)
    precisions = get_column_values(sheet, 'C', row=row, length=num_datetime_variables)
    bounds_minima = get_column_values(sheet, 'D', row=row, length=num_datetime_variables)
    bounds_maxima = get_column_values(sheet, 'E', row=row, length=num_datetime_variables)
    return format_strings, precisions, bounds_minima, bounds_maxima


def get_categorical_variables(sheet, num_categorical_variables):
    def get_a_categorical_variable():
        global row, col
        codes = []
        definitions = []
        while True:
            cell = sheet[f'{col}{row}']
            if cell.value is None:
                break
            codes.append(cell.value)
            cell = sheet[f'{next_spreadsheet_column(col)}{row}']
            definitions.append(cell.value)
            row += 1
        return codes, definitions

    global row, col
    codes_list = []
    definitions_list = []
    col = 'A'
    start_row = row + 1
    for i in range(num_categorical_variables):
        row = start_row
        codes, definitions = get_a_categorical_variable()
        codes_list.append(codes)
        definitions_list.append(definitions)
        col = next_spreadsheet_column(col, 3)
    return codes_list, definitions_list


def ingest_data_table_spreadsheet(filepath, dt_node_id):
    global row

    def set_child_node(child_name, parent_node, content=None, attribute=None):
        # Find child node.
        # If content, create child if it doesn't exist, otherwise, replace it. Set its content and, optionally, add an
        #  attribute, return the child node.
        # If no content, remove the child node if it exists. I.e., calling with content=None is a way to remove the
        #  child node, dealing with the possibility that it doesn't exist.
        child_node = parent_node.find_child(child_name)
        try:
            # We remove the child node so we can replace it with an empty one. This way, any existing attributes
            #  are removed.
            parent_node.remove_child(child_node)
        except ValueError:
            pass
        if content:
            # Note: if content is None but attribute has a value, we don't create the child node. That's intentional.
            child_node = new_child_node(child_name, parent_node, content=content, attribute=attribute)
            return child_node

    def set_bounds(domain_node, bounds_min, bounds_max):
        # Handle Bounds. Create bounds node as needed. If bounds_min or bounds_max, set the minimum or maximum,
        #  respectively, and add an attribute to the minimum to indicate that it is exclusive. If bounds_min and
        #  bounds_max are both None, remove the bounds node.
        bounds_node = domain_node.find_child(names.BOUNDS)
        if bounds_node is None and (bounds_min or bounds_max):
            bounds_node = new_child_node(names.BOUNDS, domain_node)
        if bounds_min or bounds_max:
            set_child_node(names.MINIMUM, bounds_node, bounds_min, attribute=('exclusive', 'false'))
            set_child_node(names.MAXIMUM, bounds_node, bounds_max, attribute=('exclusive', 'false'))
        elif bounds_node:
            try:
                domain_node.remove_child(bounds_node)
            except ValueError:
                pass

    def set_missing_value(attribute_node, which_mvc, mvc, mvc_explanation):
        missing_value_code_nodes = attribute_node.find_all_children(names.MISSINGVALUECODE)
        missing_value_code_node = missing_value_code_nodes[which_mvc] if which_mvc < len(missing_value_code_nodes) else None
        if mvc:
            if missing_value_code_node is None:
                missing_value_code_node = new_child_node(names.MISSINGVALUECODE, attribute_node)
            set_child_node(names.CODE, missing_value_code_node, content=mvc)
            set_child_node(names.CODEEXPLANATION, missing_value_code_node, content=mvc_explanation)
        elif missing_value_code_node:
            attribute_node.remove_child(missing_value_code_node)

    def set_numerical_variable(attribute_node, number_type, standard_unit,
                               custom_unit, custom_unit_description,
                               precision, bounds_min, bounds_max):
        if not is_numerical(attribute_node):
            return
        ir_node = interval_or_ratio_node(attribute_node)
        numeric_domain_node = ir_node.find_child(names.NUMERICDOMAIN)
        if numeric_domain_node is None:
            numeric_domain_node = new_child_node(names.NUMERICDOMAIN, ir_node)

        # Number type
        set_child_node(names.NUMBERTYPE, numeric_domain_node, number_type)

        # Precision
        set_child_node(names.PRECISION, ir_node, precision)

        # Bounds
        set_bounds(numeric_domain_node, bounds_min, bounds_max)

        # Units
        try:
            ir_node.remove_child(ir_node.find_child(names.UNIT))
        except ValueError:
            pass
        if standard_unit and standard_unit not in standard_units:
            msg = f"'{standard_unit}' is not an allowed Standard Unit. See the list of allowed Standard Units in Sheet2 of the spreadsheet. Please correct the spreadsheet and try again."
            raise exceptions.UnitIsNotAnAllowedStandardUnit(msg)
        if standard_unit or custom_unit:
            unit_node = new_child_node(names.UNIT, ir_node)
            set_child_node(names.STANDARDUNIT, unit_node, standard_unit)
            set_child_node(names.CUSTOMUNIT, unit_node, custom_unit)
            if custom_unit and not standard_unit and custom_unit_description:
                load_and_save.handle_custom_unit_additional_metadata(eml_node, custom_unit, custom_unit_description)

    def set_datetime_variable(attribute_node, format_string, precision, bounds_min, bounds_max):
        if not is_datetime(attribute_node):
            return
        datetime_node = attribute_node.find_descendant(names.DATETIME)

        # Format string
        set_child_node(names.FORMATSTRING, datetime_node, format_string)

        # Precision
        set_child_node(names.DATETIMEPRECISION, datetime_node, precision)

        # Bounds
        date_time_domain_node = attribute_node.find_descendant(names.DATETIMEDOMAIN)
        if bounds_min or bounds_max:
            if not date_time_domain_node:
                date_time_domain_node = new_child_node(names.DATETIMEDOMAIN, datetime_node)
            set_bounds(date_time_domain_node, bounds_min, bounds_max)
        elif date_time_domain_node:
            datetime_node.remove_child(date_time_domain_node)

    def set_categorical_variable(attribute_node, codes, definitions):
        if not is_categorical(attribute_node):
            return
        # Remove all codeDefinition children of enumeratedDomain
        # We'll just remove the enumeratedDomain node and recreate it
        non_numeric_domain_node = attribute_node.find_descendant(names.NONNUMERICDOMAIN)
        enumerated_domain_node = non_numeric_domain_node.find_child(names.ENUMERATEDDOMAIN)
        if enumerated_domain_node is not None:
            non_numeric_domain_node.remove_child(enumerated_domain_node)
        enumerated_domain_node = new_child_node(names.ENUMERATEDDOMAIN, non_numeric_domain_node)
        # Create the codeDefinition children
        for code, definition in zip(codes, definitions):
            code_definition_node = new_child_node(names.CODEDEFINITION, enumerated_domain_node)
            code_node = new_child_node(names.CODE, code_definition_node, content=code)
            definition_node = new_child_node(names.DEFINITION, code_definition_node, content=definition)

    current_document = current_user.get_filename()

    wb = None
    try:
        wb = load_workbook(filepath)
    except Exception as e:
        print(e)
        raise
    sheet = wb['Sheet']

    package_name = get_package_name(sheet)
    data_table_name = get_data_table_name(sheet)

    eml_node = load_eml(current_document)
    data_table_node = Node.get_node_instance(dt_node_id)
    if data_table_node is None:
        msg = 'Data table node not found. This probably indicates a software bug. Please report this issue.'
        raise exceptions.DataTableSpreadsheetError(msg)
    entity_name_node = data_table_node.find_child(names.ENTITYNAME)
    if entity_name_node is None:
        msg = 'Entity name node not found. This probably indicates a software bug. Please report this issue.'
        raise exceptions.DataTableSpreadsheetError(msg)
    entity_name = entity_name_node.content
    if entity_name != data_table_name:
        msg = f'Data table name in the spreadsheet ({data_table_name}) does not match the name in the data table metadata ({entity_name}). ' \
            'Make sure you are opening the correct spreadsheet file. Data table name changes must be performed in the ezEML editor, '\
            'not in a spreadsheet.'
        raise exceptions.DataTableNameMismatch(msg)

    column_names = get_column_names(sheet)
    length = len(column_names)
    row = length + 8
    column_types = get_column_types(sheet, length=length)
    column_definitions = get_column_definitions(sheet, length=length)
    column_labels = get_column_labels(sheet, length=length)
    storage_types = get_storage_types(sheet, length=length)
    storage_type_systems = get_storage_type_systems(sheet, length=length)
    mvc1, mvc_explanations_1, mvc2, mvc_explanations_2, mvc3, mvc_explanations_3 = get_missing_values(sheet, length=length)
    row += 4

    num_numerical_variables = column_types.count('Numerical')
    if num_numerical_variables > 0:
        number_types, standard_units_found, custom_units_found, custom_unit_descriptions, num_precisions, num_bounds_minima, num_bounds_maxima = \
            get_numerical_variables(sheet, num_numerical_variables=num_numerical_variables)
        row += num_numerical_variables + 4

    num_datetime_variables = column_types.count('DateTime')
    if num_datetime_variables > 0:
        format_strings, dt_precisions, dt_bounds_minima, dt_bounds_maxima = get_datetime_variables(sheet, num_datetime_variables=num_datetime_variables)
        row += num_datetime_variables + 4

    num_categorical_variables = column_types.count('Categorical')
    if num_categorical_variables > 0:
        codes, code_definitions = get_categorical_variables(sheet, num_categorical_variables=num_categorical_variables)

    # Check column names
    attribute_list_node = data_table_node.find_child(names.ATTRIBUTELIST)
    attribute_nodes = attribute_list_node.children
    for i, attribute_node in enumerate(attribute_nodes):
        attribute_name_node = attribute_node.find_child(names.ATTRIBUTENAME)
        if attribute_name_node.content != column_names[i]:
            msg = f'Column name "{column_names[i]}" was found in the spreadsheet where column name "{attribute_name_node.content}" was expected.<br> ' \
                    f'Column name changes must be done in the ezEML editor, not via a spreadsheet.'
            raise exceptions.DataTableNameNotFound(msg)

    # Check column types
    for i, attribute_node in enumerate(attribute_nodes):
        if compose_attribute_mscale(attribute_node) != column_types[i]:
            attribute_name_node = attribute_node.find_child(names.ATTRIBUTENAME)
            attribute_name = attribute_name_node.content
            msg = f'Column "{attribute_name}" has type "{compose_attribute_mscale(attribute_node)}" in the metadata, but ' \
                    f'type "{column_types[i]}" was found in the spreadsheet.\n Changes to column types must be performed in the ezEML ' \
                    f'editor, not via a spreadsheet.'
            raise exceptions.ColumnTypeMismatch(msg)

    # Set column definitions
    for i, attribute_node in enumerate(attribute_nodes):
        set_child_node(names.ATTRIBUTEDEFINITION, attribute_node, column_definitions[i])

    # Set column labels
    for i, attribute_node in enumerate(attribute_nodes):
        set_child_node(names.ATTRIBUTELABEL, attribute_node, column_labels[i])

    # Set storage type info
    for i, attribute_node in enumerate(attribute_nodes):
        set_child_node(names.STORAGETYPE, attribute_node, content=storage_types[i], attribute=('typeSystem', storage_type_systems[i]))

    # Set missing value codes
    # mvc1, mvc_explanations_1, mvc2, mvc_explanations_2, mvc3, mvc_explanations_3 = get_missing_values(sheet, length=length)
    for i, attribute_node in enumerate(attribute_nodes):
        set_missing_value(attribute_node, 0, mvc1[i], mvc_explanations_1[i])
        set_missing_value(attribute_node, 1, mvc2[i], mvc_explanations_2[i])
        set_missing_value(attribute_node, 2, mvc3[i], mvc_explanations_3[i])

    # Handle numerical variables
    i = 0
    for attribute_node in attribute_nodes:
        if not is_numerical(attribute_node):
            continue
        set_numerical_variable(attribute_node, number_types[i], standard_units_found[i], custom_units_found[i], custom_unit_descriptions[i],
                                   num_precisions[i], num_bounds_minima[i], num_bounds_maxima[i])
        i += 1

    # Handle datetime variables
    i = 0
    for attribute_node in attribute_nodes:
        if not is_datetime(attribute_node):
            continue
        set_datetime_variable(attribute_node, format_strings[i], dt_precisions[i], dt_bounds_minima[i], dt_bounds_maxima[i])
        i += 1

    # Handle categorical variables
    i = 0
    for attribute_node in attribute_nodes:
        if not is_categorical(attribute_node):
            continue
        set_categorical_variable(attribute_node, codes[i], code_definitions[i])
        i += 1

    save_eml(current_document, eml_node)
    get_check_metadata_status(eml_node, current_document)

    return
